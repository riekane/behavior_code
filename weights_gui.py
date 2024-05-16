import os
from tkinter import *
import time
from os import walk
import pandas as pd
from csv import DictReader, reader
import numpy as np
from datetime import date
import datetime
from upload_to_pi import reset_time, ping_host
from user_info import get_user_info
import shutil
import json
import tkinter as tk
import tkinter.font as font
from threading import Thread, Event
from functools import partial
from openpyxl import load_workbook
import fnmatch
import serial

pastel_colors = ['#ffffcc', '#99ccff', '#cc99ff', '#ff99cc', '#ffcc99', '#ffffcc', '#99ffcc', '#ccffff', '#ccccff',
                 '#ffccff', '#ffcccc', '#D3D3D3']
str_fmt = "%Y-%m-%d"


def get_today_string():
    return datetime.datetime.today().strftime(str_fmt)


def date_from_string(date_string):
    return datetime.datetime.strptime(date_string, str_fmt)


def string_from_date(datetime_obj):
    return datetime_obj.strftime(str_fmt)


def get_active_mice():
    return [
        ['ES045', 'ES046', 'ES047'],
        ['ES058', 'ES059'],
    ]


def load_log():
    with open('weight_log.json', 'r') as file:
        data = json.load(file)
    return data


def save_log(log):
    log = dict(sorted(log.items()))
    with open('weight_log.json', 'w') as file:
        json.dump(log, file, indent=2)


def make_log():
    save_log({})


def save_excel_to_log():
    log = load_log()
    wb = load_workbook('log.xlsx', read_only=True, keep_links=False)
    page_names = wb.sheetnames
    for i, page in enumerate(page_names):
        sheet = wb.worksheets[i].values
        sheet = pd.DataFrame(sheet)
        columns = sheet.iloc[0]  # grab the first row for the header
        sheet = sheet[1:]  # take the data less the header row
        sheet.columns = columns
        columns = [str(val) for val in columns.tolist()]
        if 'date' in sheet.keys():
            sheet['date'] = [d_time.strftime(str_fmt) if d_time is not None else None for d_time in sheet['date']]
        if fnmatch.fnmatch(page, 'ES***-**'):
            mice = fnmatch.filter(columns, 'ES***')
            for mouse in mice:
                mouse_weights = sheet[['date', mouse]].copy()
                mouse_weights = mouse_weights.dropna()
                mouse_weights[mouse] = [[val] for val in mouse_weights[mouse]]
                mouse_dict = dict(zip(mouse_weights['date'], mouse_weights[mouse]))
                if mouse in log.keys():
                    for key in log[mouse].keys():
                        if key in mouse_dict.keys():
                            log[mouse][key] = np.unique(np.array(log[mouse][key] + mouse_dict[key])).tolist()
                    for key in mouse_dict.keys():
                        if key not in log[mouse].keys():
                            log[mouse][key] = mouse_dict[key]
                else:
                    log[mouse] = mouse_dict
        elif fnmatch.fnmatch(page, 'ES***'):
            mouse = page
            if 'weight' not in sheet.keys():
                print()
            mouse_weights = sheet[['date', 'weight']].copy()
            mouse_weights = mouse_weights.dropna()
            mouse_weights['weight'] = [[val] for val in mouse_weights['weight']]
            mouse_dict = dict(zip(mouse_weights['date'], mouse_weights['weight']))
            if mouse in log.keys():
                for key in log[mouse].keys():
                    if key in mouse_dict.keys():
                        log[mouse][key] = np.unique(np.array(log[mouse][key] + mouse_dict[key])).tolist()
                for key in mouse_dict.keys():
                    if key not in log[mouse].keys():
                        log[mouse][key] = mouse_dict[key]
            else:
                log[mouse] = mouse_dict
    save_log(log)


class Scale:
    def __init__(self, check_tare=False):
        self.port = 'COM3'
        self.BAUD = 9600
        self.PARITY = serial.PARITY_ODD
        self.STOP_BITS = serial.STOPBITS_ONE
        self.BYTE_SIZE = serial.SIXBITS
        self.canceled = False
        self.ser = serial.Serial(self.port, self.BAUD, self.BYTE_SIZE, self.PARITY, self.STOP_BITS)

        self.root = tk.Tk()
        self.root.geometry("300x200+2500+25")
        self.root.title('Cancel Button')
        self.button = tk.Button(
            master=self.root,
            text='Cancel\nWeighing',
            width=50,
            height=10,
            bg="white",
            fg="black",
            font=("Arial", 25),
            command=self.cancel)
        self.button.pack()
        self._job = None
        self.weight = None
        self.zeroed = False
        self.weight_history = []
        self.check_tare = check_tare

    def cleanse(self):
        try:
            # self.ser = serial.Serial(self.port, self.BAUD, self.BYTE_SIZE, self.PARITY, self.STOP_BITS)
            # temp = serial.Serial(port, baud, bytesize, parity, stopbits)

            # Flush both input/output buffer.
            self.ser.reset_input_buffer()
            self.ser.reset_output_buffer()

            # self.ser.close()
        except:
            print('Unable to open/clean COM port')
            return

    def get_weight(self):
        line = ''
        self.ser.reset_input_buffer()
        purge = True
        while True:
            if self.ser.in_waiting > 0:
                x = self.ser.read(1).decode()
                if len(x) and x not in [' ', '\'']:
                    line += x
                if line[-1:] == '\n':
                    if purge:
                        line = ''
                        purge = False
                    else:
                        w = float(line[:-1])
                        break
        return w

    def weigh_logic(self):
        tic = time.time()
        w = self.get_weight()

        if not self.zeroed and -.2 < w < .2 and self.check_tare:
            self.weight_history.append(w)
            if len(self.weight_history) > 8 and np.std(np.array(self.weight_history)) < .1 and abs(
                    np.mean(np.array(self.weight_history))) < .1:
                self.zeroed = True
                self.weight_history = []
            if len(self.weight_history) > 10:
                print(
                    f'{np.mean(np.array(self.weight_history)):.2f} +- {np.std(np.array(self.weight_history)):.2f} != 0')
                self.weight_history = self.weight_history[-10:]

        if not self.zeroed and not self.check_tare and w < 0:
            self.zeroed = True

        if self.zeroed and 10 < w < 45:
            self.weight_history.append(w)
            if len(self.weight_history) > 25 and np.std(np.array(self.weight_history)) < .5:
                self.weight = np.median(np.array(self.weight_history))
                print(f'final weight: {self.weight}')
                self.end()
            if len(self.weight_history) > 30:
                print(
                    f'{np.mean(np.array(self.weight_history)):.2f} + or - {np.std(np.array(self.weight_history)):.2f}')
                self.weight_history = self.weight_history[-30:]
        t_remaining = .1 - (time.time() - tic)
        if t_remaining > 0:
            self._job = self.root.after(round(t_remaining * 1000), self.weigh_logic)
        else:
            self._job = self.root.after(5, self.weigh_logic)

    def weigh_one(self):
        self.cleanse()
        self._job = self.root.after(10, self.weigh_logic)
        self.root.mainloop()
        self.root.destroy()
        if self.canceled:
            self.canceled = False
            return 'canceled'
        else:
            return self.weight

    def end(self):
        self.ser.close()
        if self._job is not None:
            self.root.after_cancel(self._job)
            self._job = None
        self.root.quit()

    def cancel(self):
        self.canceled = True
        self.end()


#     def weigh_one(self, check_tare=True):
#         stop_event = Event()
#         t1 = Thread(target=self.cancel_button, args=[stop_event])
#         # t1.setDaemon(True)
#         t1.start()
#         self.cleanse()
#         zeroed = False
#         weight_history = []
#         with serial.Serial(self.port, self.BAUD, self.BYTE_SIZE, self.PARITY, self.STOP_BITS) as ser:
#             while True:
#                 tic = time.time()
#                 if self.cancel:
#                     break
#                 w = self.get_weight(ser)
#
#                 if not zeroed and -.2 < w < .2 and check_tare:
#                     weight_history.append(w)
#                     if len(weight_history) > 8 and np.std(np.array(weight_history)) < .1 and abs(
#                             np.mean(np.array(weight_history))) < .1:
#                         zeroed = True
#                         weight_history = []
#                     if len(weight_history) > 10:
#                         print(f'{np.mean(np.array(weight_history)):.2f} +- {np.std(np.array(weight_history)):.2f} != 0')
#                         weight_history = weight_history[-10:]
#
#                 if not zeroed and not check_tare and w < 0:
#                     zeroed = True
#
#                 if zeroed and 15 < w < 45:
#                     weight_history.append(w)
#                     if len(weight_history) > 25 and np.std(np.array(weight_history)) < .5:
#                         weight = np.median(np.array(weight_history))
#                         print(f'final weight: {weight}')
#                         break
#                     if len(weight_history) > 30:
#                         print(f'{np.mean(np.array(weight_history)):.2f} + or - {np.std(np.array(weight_history)):.2f}')
#                         weight_history = weight_history[-30:]
#                 t_remaining = .1 - (time.time() - tic)
#                 if t_remaining > 0:
#                     time.sleep(t_remaining)
#                 # print(f'loop time: {time.time()-tic:.2f}')
#         stop_event.set()
#         print('left over thread?')
#         print(t1.is_alive())
#         if self.cancel:
#             self.cancel = False
#             return 'canceled'
#         else:
#             return weight
#
#     def cancel_button(self, stop_event):
#         self.app = CancelButton(self, stop_event)
#
#
# class CancelButton:
#     def __init__(self, scale_instance, stop_event):
#         self.root = tk.Tk()
#         self.root.geometry("300x200+2500+25")
#         self.root.title('Cancel Button')
#         self.scale_instance = scale_instance
#         self.stop_event = stop_event
#         self.button = tk.Button(
#             master=self.root,
#             text='Cancel\nWeighing',
#             width=50,
#             height=10,
#             bg="white",
#             fg="black",
#             font=("Arial", 25),
#             command=self.stop)
#         self.button.pack()
#         self._job = self.root.after(1000, self.check_continue)
#         self.root.mainloop()
#
#     def stop(self):
#         self.scale_instance.cancel = True
#         if self._job is not None:
#             self.root.after_cancel(self._job)
#             self._job = None
#         self.root.destroy()
#
#     def check_continue(self):
#         if self.stop_event.is_set():
#             if self._job is not None:
#                 self.root.after_cancel(self._job)
#                 self._job = None
#             self.root.destroy()
#         else:
#             self._job = self.root.after(1000, self.check_continue)


def test_scale2():
    port = 'COM3'
    BAUD = 9600
    PARITY = serial.PARITY_ODD
    BYTESIZE = serial.SIXBITS
    STOPBITS = serial.STOPBITS_ONE
    try:
        temp = serial.Serial(port, BAUD, BYTESIZE, PARITY, STOPBITS)
        # temp = serial.Serial(port, baud, bytesize, parity, stopbits)

        # Flush both input/output buffer.
        temp.reset_input_buffer()
        temp.reset_output_buffer()

        temp.close()
    except:
        print('Unable to open/clean ' + port + ':' + str(BAUD) + ',' + str(PARITY))
        return

    t = time.time()
    line = ''
    with serial.Serial(port, BAUD, BYTESIZE, PARITY, STOPBITS, timeout=.5) as ser:
        while time.time() < t + 60:
            if ser.in_waiting > 0:
                x = ser.read(1).decode()
                if len(x) and x not in [' ', '\'']:
                    line += x
                if line[-1:] == '\n':
                    print(line, end='')
                    # print(x, end='')
                    # char_num += 1
            else:
                time.sleep(.1)


def test_scale():
    baud_rates = [9600]
    # baud_rates = [110, 300, 600, 1200, 2400, 4800, 9600, 14400, 19200, 38400, 57600, 115200, 128000, 256000]
    parities = [serial.PARITY_ODD, serial.PARITY_EVEN]
    characters = [''.join([chr(i), chr(13)]) for i in list(range(97, 123))]
    # characters = list(range(65, 91)) + list(range(97, 123))
    # characters = [''.join([chr(i), chr(13)]) for i in characters] + [chr(i) for i in characters]
    # characters = ['weight\r']
    # bytesizes = [serial.EIGHTBITS]
    stop_bits = [serial.STOPBITS_ONE]
    bytesizes = [serial.SEVENBITS]
    # bytesizes = [serial.FIVEBITS, serial.SIXBITS, serial.SEVENBITS, serial.EIGHTBITS]
    # stop_bits = [serial.STOPBITS_ONE, serial.STOPBITS_TWO, serial.STOPBITS_ONE_POINT_FIVE]
    port = 'COM3'
    # port = '/dev/ttyUSB0'
    # baudrate = 9600
    # parity = serial.PARITY_NONE
    # stopbits = serial.STOPBITS_ONE
    # bytesize = serial.EIGHTBITS
    for BAUD in baud_rates:
        for PARITY in parities:
            for BYTESIZE in bytesizes:
                for STOPBITS in stop_bits:
                    print(f'baud: {BAUD}, parity: {PARITY}, bytesize: {BYTESIZE}, stopbits: {STOPBITS}: ', end='')
                    success = False
                    try:
                        temp = serial.Serial(port, BAUD, BYTESIZE, PARITY, STOPBITS)
                        # temp = serial.Serial(port, baud, bytesize, parity, stopbits)

                        # Flush both input/output buffer.
                        temp.reset_input_buffer()
                        temp.reset_output_buffer()

                        temp.close()
                    except:
                        print('Unable to open/clean ' + port + ':' + str(BAUD) + ',' + str(PARITY))
                        return

                    # You are using Ctrl+C to stop the program. Using *with* ensures
                    # that the serial port is closed when you exit the program.
                    with serial.Serial(port, BAUD, BYTESIZE, PARITY, STOPBITS, timeout=.5) as ser:
                        if ser.in_waiting > 0:
                            x = ser.read(10)
                            print(f'no input: {x}')
                            success = True
                        for char in characters:
                            # ser.write(char.encode())
                            time.sleep(.1)
                            # print(char)
                            if ser.in_waiting > 0:
                                x = ser.read(10)
                                print(f'{char[0]}: {x}')
                                success = True
                    if success:
                        print(f'baud: {BAUD}, parity: {PARITY}, bytesize: {BYTESIZE}, stopbits: {STOPBITS}: success!')
                    else:
                        print('unsuccessful')

                # for _ in range(1000):
                #     # Only read data if there are bytes already waiting in the buffer.
                #     if ser.in_waiting <= 0:
                #         time.sleep(0.1)
                #         continue
                #
                #     # We got bytes, read them.
                #     x = ser.readline()
                #     print(x)
                #     time.sleep(1)


def make_button(frame, f, text, color, button_font):
    return tk.Button(
        text=text,
        font=button_font,
        width=10,
        height=5,
        bg=color,
        fg="black",
        master=frame,
        command=f)


def make_display(frame):
    return Label(frame, text="", fg="Black", font=("Helvetica", 18))


class App:
    def __init__(self, master=None):
        self.root = tk.Tk()
        self.root.geometry('500x1000+1950+25')
        self.root.title('Weights Gui')
        self.font = font.Font(size=15)
        self.log = load_log()
        self.today = str(date.today())

        # Frame.__init__(self, master)
        # self.master = master
        self.mice = get_active_mice()
        self.mouse_names = [item for sublist in self.mice for item in sublist]

        self.from_here_button_list = []
        self.from_here_functions = [partial(self.weigh_all, i) for i in range(len(self.mouse_names))]
        self.root.rowconfigure(0, weight=1, minsize=50)
        self.root.columnconfigure(0, weight=1, minsize=5)
        button_i = 1
        for i, cage in enumerate(self.mice):
            for j, mouse in enumerate(cage):
                self.root.rowconfigure(button_i, weight=1, minsize=30)
                frame = tk.Frame(master=self.root, borderwidth=1)
                frame.grid(row=button_i, column=0, sticky="nsew")
                button = make_button(frame, self.from_here_functions[button_i - 1], '\\/', 'lightgray', self.font)
                button.pack(fill=tk.BOTH, expand=True)
                self.from_here_button_list.append(button)
                button_i += 1

        self.button_list = []
        self.functions = [partial(self.record_weight, i) for i in range(len(self.mouse_names))]
        self.root.rowconfigure(0, weight=1, minsize=50)
        self.root.columnconfigure(1, weight=10, minsize=75)
        button_i = 1
        for i, cage in enumerate(self.mice):
            for j, mouse in enumerate(cage):
                self.root.rowconfigure(button_i, weight=1, minsize=30)
                frame = tk.Frame(master=self.root, borderwidth=1)
                frame.grid(row=button_i, column=1, sticky="nsew")
                button = make_button(frame, self.functions[button_i - 1], mouse, pastel_colors[i], self.font)
                button.pack(fill=tk.BOTH, expand=True)
                self.button_list.append(button)
                button_i += 1
        self.root.rowconfigure(button_i, weight=1, minsize=30)
        frame = tk.Frame(master=self.root, borderwidth=1)
        frame.grid(row=button_i, column=1, sticky="nsew")
        button = make_button(frame, partial(self.weigh_all, 0), 'Weigh All', 'white', self.font)
        button.pack(fill=tk.BOTH, expand=True)

        self.weight_label_list = []
        self.root.columnconfigure(2, weight=4, minsize=75)
        for i in range(len(self.mouse_names)):
            if self.mouse_names[i] not in self.log.keys():
                self.log[self.mouse_names[i]] = {}

            frame = tk.Frame(master=self.root, borderwidth=1)
            frame.grid(row=i + 1, column=2, sticky="nsew")
            label = make_display(frame)
            label.pack(fill=tk.BOTH, expand=True)
            today = self.get_today(self.mouse_names[i])
            recent = self.get_recent(self.mouse_names[i])
            if today is not None:
                label.configure(text=f'{today:.1f}g', fg='black')
            elif recent is not None:
                label.configure(text=f'{recent:.1f}g', fg='gray')
            self.weight_label_list.append(label)

        self.percent_label_list = []
        self.root.columnconfigure(3, weight=4, minsize=75)
        for i in range(len(self.mouse_names)):
            frame = tk.Frame(master=self.root, borderwidth=1)
            frame.grid(row=i + 1, column=3, sticky="nsew")
            label = make_display(frame)
            label.pack(fill=tk.BOTH, expand=True)
            percent = self.get_percent(self.mouse_names[i])
            if percent is not None:
                if percent > 85:
                    label.configure(text=f'{percent:.1f}%', fg='black')
                elif percent <= 85 and percent >= 80:
                    label.configure(text=f'{percent:.1f}%', fg='orange')
                elif percent < 80:
                    label.configure(text=f'{percent:.1f}%', fg='red')
            self.percent_label_list.append(label)
        headers = ['Mouse', 'Weight', 'Percent']
        for i, header in enumerate(headers):
            frame = tk.Frame(master=self.root, borderwidth=1)
            frame.grid(row=0, column=i + 1, sticky="nsew")
            label = make_display(frame)
            label.pack(fill=tk.BOTH, expand=True)
            label.configure(text=header, fg='black')
        self._job = self.root.after(int(28000e3), self.update_all)
        self.root.mainloop()

    def update_all(self):
        if self.today != str(date.today()):
            self.today = str(date.today())
            for label in self.weight_label_list:
                label.configure(fg='gray')
            self.button_list[0].update()
            self._job = self.root.after(int(80000e3), self.update_all)
        else:
            self._job = self.root.after(int(3600e3), self.update_all)

    def record_weight(self, button_i, check_tare=False):
        today = get_today_string()
        print(self.mouse_names[button_i])
        scale = Scale(check_tare=check_tare)
        weight = scale.weigh_one()
        if weight != 'canceled' and weight is not None:
            if today in self.log[self.mouse_names[button_i]].keys():
                self.log[self.mouse_names[button_i]][today].append(weight)
            else:
                self.log[self.mouse_names[button_i]][today] = [weight]
            save_log(self.log)
            self.update_display(button_i)
        return weight

    def weigh_all(self, start=0):
        k = 0
        result = ''
        for i, cage in enumerate(self.mice):
            for _ in cage:
                if k < start:
                    k += 1
                    continue
                self.button_list[k].configure(bg='blue')
                self.button_list[0].update()
                result = self.record_weight(k, check_tare=k == start)
                self.button_list[k].configure(bg=pastel_colors[i])
                self.button_list[0].update()
                k += 1
                if result == 'canceled' or result is None:
                    break
            if result == 'canceled' or result is None:
                print('Weighing canceled')
                break

    def update_display(self, i):
        today = self.get_today(self.mouse_names[i])
        self.weight_label_list[i].configure(text=f'{today:.1f}g', fg='black')

        percent = self.get_percent(self.mouse_names[i])
        if percent is not None:
            if percent > 85:
                self.percent_label_list[i].configure(text=f'{percent:.1f}%', fg='black')
            elif percent <= 85 and percent >= 80:
                self.percent_label_list[i].configure(text=f'{percent:.1f}%', fg='orange')
            elif percent < 80:
                self.percent_label_list[i].configure(text=f'{percent:.1f}%', fg='red')

    def get_today(self, mouse):
        today = get_today_string()
        if today in self.log[mouse].keys():
            return self.log[mouse][today][-1]
        else:
            return None

    def get_recent(self, mouse):
        if len(self.log[mouse].keys()):
            keys = list(self.log[mouse].keys())
            keys.sort()
            return self.log[mouse][keys[-1]][-1]
        else:
            return None

    def get_percent(self, mouse):
        if len(self.log[mouse].keys()):
            keys = list(self.log[mouse].keys())
            keys.sort()
            current = self.log[mouse][keys[-1]][-1]
            max_weight = np.max([val for sublist in self.log[mouse].values() for val in sublist])
            return current / max_weight * 100
        else:
            return None


# def run_gui():
#     root = Tk()
#     app = App(root)
#     root.wm_title("Weights Gui")
#     root.geometry("600x400")
#     root.mainloop()

def run_gui():
    app = App()


if __name__ == '__main__':
    run_gui()
    # save_excel_to_log()
    # test_scale2()
